"""Read-only XLSX-to-HTML preview rendering for the web review flow."""

from __future__ import annotations

import base64
import html
import io
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter, range_boundaries


def render_workbook_preview(
    workbook_path: Path,
    output_path: Path,
    *,
    title: str,
) -> Path:
    formulas = load_workbook(workbook_path, data_only=False)
    values = load_workbook(workbook_path, data_only=True)
    worksheet = formulas.active
    value_sheet = values[worksheet.title]
    max_column = _preview_max_column(worksheet)
    striped_cells = _striped_table_cells(worksheet)
    merged = {
        (cell.row, cell.column): range_
        for range_ in worksheet.merged_cells.ranges
        for cell in (worksheet.cell(range_.min_row, range_.min_col),)
    }
    merged_children = {
        (row, column)
        for range_ in worksheet.merged_cells.ranges
        for row in range(range_.min_row, range_.max_row + 1)
        for column in range(range_.min_col, range_.max_col + 1)
        if (row, column) != (range_.min_row, range_.min_col)
    }

    columns = []
    for column in range(1, max_column + 1):
        letter = get_column_letter(column)
        width = worksheet.column_dimensions[letter].width or 8.43
        columns.append(f'<col style="width:{_column_pixels(width):.1f}px">')

    rows: list[str] = []
    for row in range(1, worksheet.max_row + 1):
        height = worksheet.row_dimensions[row].height or 15
        cells: list[str] = []
        for column in range(1, max_column + 1):
            if (row, column) in merged_children:
                continue
            cell = worksheet.cell(row, column)
            if isinstance(cell, MergedCell):
                continue
            range_ = merged.get((row, column))
            span = ""
            if range_ is not None:
                span = (
                    f' rowspan="{range_.max_row - range_.min_row + 1}"'
                    f' colspan="{range_.max_col - range_.min_col + 1}"'
                )
            display_value = value_sheet.cell(row, column).value
            if display_value is None:
                display_value = cell.value
            value = _display(display_value, cell.number_format)
            cells.append(
                f'<td{span} style="{_cell_style(cell, (row, column) in striped_cells)}">'
                f'<span>{html.escape(value)}</span></td>'
            )
        rows.append(
            f'<tr style="height:{max(height * 1.333, 12):.1f}px">'
            + "".join(cells)
            + "</tr>"
        )

    sheet_width = sum(
        _column_pixels(
            worksheet.column_dimensions[get_column_letter(column)].width or 8.43
        )
        for column in range(1, max_column + 1)
    )
    sheet_offset = max((734 - sheet_width) / 2, 0)
    images = "".join(
        _image_html(image, worksheet, sheet_offset) for image in worksheet._images
    )
    body = f"""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{html.escape(title)}</title>
<style>
*{{box-sizing:border-box}}
html,body{{margin:0;background:#eef1ef}}
.viewport{{padding:24px;display:flex;justify-content:center;min-height:100vh}}
.page{{position:relative;background:#fff;width:794px;min-height:1123px;padding:30px;
box-shadow:0 14px 45px rgba(22,32,27,.14);overflow:hidden;transform-origin:top center}}
table{{width:{sheet_width:.1f}px;margin:0 auto;border-collapse:collapse;table-layout:fixed}}
td{{padding:0 3px;overflow:hidden;white-space:pre-wrap;word-break:keep-all;line-height:1.2}}
td span{{display:block;max-width:100%;overflow:hidden}}
.sheet-image{{position:absolute;object-fit:contain;pointer-events:none}}
@media(max-width:850px){{.viewport{{padding:10px}}.page{{transform:scale(calc((100vw - 20px)/794));margin-bottom:calc(-1123px + (100vw - 20px)*1.414)}}}}
</style>
</head>
<body><main class="viewport"><article class="page">{images}<table><colgroup>{''.join(columns)}</colgroup>
<tbody>{''.join(rows)}</tbody></table></article></main></body></html>"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(body, encoding="utf-8")
    formulas.close()
    values.close()
    return output_path


def _display(value: Any, number_format: str) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        if any(token in number_format for token in ("#,##0", "#,###")):
            return f"{value:,.0f}"
        return str(value)
    return str(value)


def _cell_style(cell: Any, table_stripe: bool = False) -> str:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    declarations = [
        f"font-family:{_font_family(font.name)}",
        f"font-size:{font.sz or 10}pt",
        f"font-weight:{'700' if font.bold else '400'}",
        f"font-style:{'italic' if font.italic else 'normal'}",
        f"text-align:{alignment.horizontal or 'left'}",
        f"vertical-align:{alignment.vertical or 'middle'}",
    ]
    font_color = _color(font.color)
    if font_color:
        declarations.append(f"color:{font_color}")
    fill_color = _color(fill.fgColor)
    if fill.fill_type and fill_color:
        declarations.append(f"background:{fill_color}")
    elif table_stripe:
        declarations.append("background:#F8F5EE")
    if alignment.wrap_text:
        declarations.append("white-space:pre-wrap")
    side_names = (("top", border.top), ("right", border.right), ("bottom", border.bottom), ("left", border.left))
    for name, side in side_names:
        if side is not None and side.style:
            color = _color(side.color) or "#303330"
            width = "2px" if side.style in {"medium", "thick", "double"} else "1px"
            declarations.append(f"border-{name}:{width} solid {color}")
    return ";".join(declarations)


def _color(color: Color | None) -> str | None:
    if color is None or color.type != "rgb" or not color.rgb:
        return None
    value = str(color.rgb)
    return f"#{value[-6:]}"


def _image_html(image: Any, worksheet: Any, sheet_offset: float = 0) -> str:
    anchor = image.anchor
    if not hasattr(anchor, "_from"):
        return ""
    start = anchor._from
    left = 30 + sheet_offset + sum(
        _column_pixels(
            worksheet.column_dimensions[get_column_letter(column)].width or 8.43
        )
        for column in range(1, start.col + 1)
    )
    left += start.colOff / 9525
    top = 30 + sum(
        (worksheet.row_dimensions[row].height or 15) * 1.333
        for row in range(1, start.row + 1)
    )
    top += start.rowOff / 9525
    try:
        data = image._data()
    except Exception:
        return ""
    extension = getattr(image, "format", "png").lower()
    encoded = base64.b64encode(data).decode("ascii")
    return (
        f'<img class="sheet-image" alt="" src="data:image/{extension};base64,{encoded}" '
        f'style="left:{left:.1f}px;top:{top:.1f}px;width:{image.width}px;height:{image.height}px">'
    )


def _column_pixels(width: float) -> float:
    """Match Excel's character-width to pixel conversion for table geometry."""
    if width < 1:
        return width * 12
    return int(((256 * width + int(128 / 7)) / 256) * 7)


def _font_family(name: str | None) -> str:
    """Declare the workbook font first; fallbacks affect preview only."""
    escaped = (name or "Arial").replace("\\", "\\\\").replace("'", "\\'")
    return f"'{escaped}','Apple SD Gothic Neo','Noto Sans KR',Arial,sans-serif"


def _preview_max_column(worksheet: Any) -> int:
    meaningful = [
        cell.column
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
        and str(cell.value).strip()
        and str(cell.value).strip() != "`"
    ]
    table_columns = [
        table.ref.split(":")[-1]
        for table in worksheet.tables.values()
    ]
    for reference in table_columns:
        letters = "".join(character for character in reference if character.isalpha())
        if letters:
            from openpyxl.utils.cell import column_index_from_string

            meaningful.append(column_index_from_string(letters))
    return max(meaningful, default=worksheet.max_column)


def _striped_table_cells(worksheet: Any) -> set[tuple[int, int]]:
    striped: set[tuple[int, int]] = set()
    for table in worksheet.tables.values():
        style = table.tableStyleInfo
        if style is None or not style.showRowStripes:
            continue
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        data_rows = [
            row
            for row in range(min_row + 1, max_row + 1)
            if worksheet.row_dimensions[row].height is not None
        ]
        for row in data_rows[::2]:
            for column in range(min_col, max_col + 1):
                striped.add((row, column))
    return striped
