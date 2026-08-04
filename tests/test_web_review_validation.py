from __future__ import annotations

import json
import subprocess
from pathlib import Path

from openpyxl import load_workbook

from document_generator import RmntcDocumentGenerator


ROOT = Path(__file__).resolve().parents[1]
VALIDATOR = ROOT / "web/lib/reviewValidation.ts"
REVIEW_COMPONENT = ROOT / "web/components/ReviewWorkspace.tsx"


def _validate(draft: dict) -> list[dict[str, str]]:
    script = (
        "import fs from 'node:fs';"
        f"import {{validateCanonicalDraft}} from {json.dumps(VALIDATOR.as_uri())};"
        "const draft=JSON.parse(fs.readFileSync(0,'utf8'));"
        "process.stdout.write(JSON.stringify(validateCanonicalDraft(draft)));"
    )
    completed = subprocess.run(
        [
            "node",
            "--disable-warning=MODULE_TYPELESS_PACKAGE_JSON",
            "--experimental-strip-types",
            "--input-type=module",
            "-e",
            script,
        ],
        cwd=ROOT,
        input=json.dumps(draft, ensure_ascii=False),
        text=True,
        capture_output=True,
        check=True,
    )
    return json.loads(completed.stdout)


def _schema_string_fields(draft: dict) -> list[dict]:
    schema_module = ROOT / "web/lib/canonicalSchema.ts"
    script = (
        "import fs from 'node:fs';"
        f"import {{canonicalStringFields}} from {json.dumps(schema_module.as_uri())};"
        "const draft=JSON.parse(fs.readFileSync(0,'utf8'));"
        "process.stdout.write(JSON.stringify(canonicalStringFields(draft)));"
    )
    completed = subprocess.run(
        [
            "node",
            "--disable-warning=MODULE_TYPELESS_PACKAGE_JSON",
            "--experimental-strip-types",
            "--input-type=module",
            "-e",
            script,
        ],
        cwd=ROOT,
        input=json.dumps(draft, ensure_ascii=False),
        text=True,
        capture_output=True,
        check=True,
    )
    return json.loads(completed.stdout)


def test_required_string_fields_are_discovered_from_canonical_schema():
    draft = json.loads((ROOT / "input/invoice.json").read_text(encoding="utf-8"))
    fields = _schema_string_fields(draft)
    required = {field["path"] for field in fields if field["required"]}

    assert required == {
        "schema_version",
        "document_type",
        "document.invoice_number",
        "document.dates.issue_date",
        "document.seller.name",
        "document.buyer.name",
        "document.currency",
        "document.items.0.description",
        "document.items.0.unit",
        "document.items.1.description",
        "document.items.1.unit",
        "document.items.2.description",
        "document.items.2.unit",
    }
    source = REVIEW_COMPONENT.read_text(encoding="utf-8")
    for path in (
        "schema_version",
        "document_type",
        "document.invoice_number",
        "document.dates.issue_date",
        "document.currency",
        "document.items.${index}.description",
        "document.items.${index}.unit",
    ):
        assert path in source
    assert 'prefix="document.seller"' in source
    assert 'prefix="document.buyer"' in source
    assert "isRequiredReviewString(draft," in source


def test_missing_item_unit_blocks_review_submit_before_api_call():
    draft = json.loads((ROOT / "input/invoice.json").read_text(encoding="utf-8"))
    del draft["document"]["items"][0]["unit"]

    issues = _validate(draft)

    assert issues == [
        {
            "path": "document.items.0.unit",
            "message": "품목 1의 단위를 입력해 주세요.",
        }
    ]
    source = REVIEW_COMPONENT.read_text(encoding="utf-8")
    local_validation = source.index(
        "const localIssues = validateCanonicalDraft(draft);"
    )
    api_call = source.index('fetch("/api/generate"')
    assert local_validation < api_call
    assert (
        "disabled={generating || requiredIssues.length > 0 || "
        "markupValidation.value === undefined}"
    ) in source
    assert 'response.status === 422' in source
    assert "setError(message)" in source


def test_entered_unit_passes_review_and_reaches_downstream_workbook(tmp_path):
    draft = json.loads((ROOT / "input/invoice.json").read_text(encoding="utf-8"))
    draft["document"]["items"][0].pop("unit")
    assert _validate(draft)

    draft["document"]["items"][0]["unit"] = "개"
    assert _validate(draft) == []

    approved = tmp_path / "approved_invoice.json"
    approved.write_text(
        json.dumps(draft, ensure_ascii=False),
        encoding="utf-8",
    )
    generated = tmp_path / "generated"
    RmntcDocumentGenerator(ROOT).generate(approved, generated)

    preserved = json.loads(approved.read_text(encoding="utf-8"))
    assert preserved["document"]["items"][0]["unit"] == "개"
    workbook = load_workbook(generated / "comparison.xlsx", data_only=False)
    assert workbook["Sheet1"]["D14"].value == "개"
    workbook.close()


def test_seller_phone_is_optional_and_manual_text_is_preserved(tmp_path):
    draft = json.loads((ROOT / "input/invoice.json").read_text(encoding="utf-8"))
    draft["document"]["seller"]["contact"].pop("phone")
    assert _validate(draft) == []

    approved = tmp_path / "approved_without_phone.json"
    approved.write_text(json.dumps(draft, ensure_ascii=False), encoding="utf-8")
    generated = tmp_path / "generated_without_phone"
    RmntcDocumentGenerator(ROOT).generate(approved, generated)
    quotation = load_workbook(generated / "quotation.xlsx", data_only=False)
    assert quotation["견적서"]["A35"].value == "E-mail : billing@rmntc.example"
    assert "미기재" not in quotation["견적서"]["A35"].value
    quotation.close()
    comparison = load_workbook(generated / "comparison.xlsx", data_only=False)
    assert comparison["Sheet1"]["F9"].value in (None, "")
    assert comparison["Sheet1"]["I9"].value in (None, "")
    comparison.close()

    draft["document"]["seller"]["contact"]["phone"] = 212345678
    assert _validate(draft) == [
        {
            "path": "document.seller.contact.phone",
            "message": "공급자 전화번호는 문자로 입력해 주세요.",
        }
    ]

    draft["document"]["seller"]["contact"]["phone"] = "02-9876-5432"
    assert _validate(draft) == []
    approved = tmp_path / "approved_invoice.json"
    approved.write_text(
        json.dumps(draft, ensure_ascii=False),
        encoding="utf-8",
    )
    generated = tmp_path / "generated"
    RmntcDocumentGenerator(ROOT).generate(approved, generated)

    preserved = json.loads(approved.read_text(encoding="utf-8"))
    assert preserved["document"]["seller"]["contact"]["phone"] == "02-9876-5432"
    quotation = load_workbook(generated / "quotation.xlsx", data_only=False)
    assert "TEL : 02-9876-5432" in quotation["견적서"]["A35"].value
    quotation.close()
    comparison = load_workbook(generated / "comparison.xlsx", data_only=False)
    assert comparison["Sheet1"]["F9"].value == "02-9876-5432"
    assert comparison["Sheet1"]["I9"].value == "02-9876-5432"
    comparison.close()


def test_optional_schema_string_rejects_numbers_without_becoming_required():
    draft = json.loads((ROOT / "input/invoice.json").read_text(encoding="utf-8"))
    draft["document"]["buyer"]["contact"].pop("phone")
    assert _validate(draft) == []

    draft["document"]["buyer"]["contact"]["phone"] = 1098765432
    assert _validate(draft) == [
        {
            "path": "document.buyer.contact.phone",
            "message": "공급받는자 전화번호는 문자로 입력해 주세요.",
        }
    ]


def test_required_schema_strings_reject_null_numbers_and_objects():
    draft = json.loads((ROOT / "input/invoice.json").read_text(encoding="utf-8"))
    draft["schema_version"] = 1
    draft["document"]["buyer"]["name"] = None
    draft["document"]["items"][0]["unit"] = {"label": "개"}

    issues = {issue["path"]: issue["message"] for issue in _validate(draft)}
    assert issues["schema_version"] == "스키마 버전은 문자로 입력해 주세요."
    assert issues["document.buyer.name"] == "공급받는자 상호를 입력해 주세요."
    assert issues["document.items.0.unit"] == "품목 1의 단위는 문자로 입력해 주세요."
