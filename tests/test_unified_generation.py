from __future__ import annotations

import hashlib
import json
import subprocess
import sys
from copy import deepcopy
from dataclasses import replace
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from document_generator import GenerationError, RmntcDocumentGenerator


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "input/invoice.json"


def _template_hashes(generator: RmntcDocumentGenerator) -> dict[str, str]:
    return {
        spec.document: hashlib.sha256(spec.template_path.read_bytes()).hexdigest()
        for spec in generator.templates
    }


def _expanded_invoice(item_count: int) -> dict:
    invoice = json.loads(SOURCE.read_text(encoding="utf-8"))
    original = invoice["document"]["items"]
    items = []
    for index in range(item_count):
        item = deepcopy(original[index % len(original)])
        item["line_number"] = index + 1
        item["description"] = f"{item['description']} {index + 1}"
        items.append(item)
    invoice["document"]["items"] = items
    invoice["document"]["totals"] = {
        field: sum(item[field] for item in items)
        for field in ("supply_amount", "vat", "total")
    }
    return invoice


def _write_invoice(path: Path, item_count: int) -> Path:
    path.write_text(
        json.dumps(_expanded_invoice(item_count), ensure_ascii=False),
        encoding="utf-8",
    )
    return path


def test_successful_generation_of_all_documents(tmp_path):
    generator = RmntcDocumentGenerator(ROOT)
    hashes = _template_hashes(generator)
    output = tmp_path / "output"

    report = generator.generate(SOURCE, output)

    assert report.company == "rmntc"
    assert report.comparison_markup == Decimal("0.08")
    assert {result.document: result.status for result in report.documents} == {
        "statement": "success",
        "quotation": "success",
        "comparison": "success",
    }
    assert {result.document: result.total for result in report.documents} == {
        "statement": Decimal("3850000"),
        "quotation": Decimal("3850000"),
        "comparison": Decimal("3780000"),
    }
    expected_sheets = {
        "statement.xlsx": "청구서",
        "quotation.xlsx": "견적서",
        "comparison.xlsx": "Sheet1",
    }
    for name, worksheet in expected_sheets.items():
        path = output / name
        assert path.is_file()
        workbook = load_workbook(path, data_only=False)
        assert worksheet in workbook.sheetnames
        workbook.close()
    assert _template_hashes(generator) == hashes


def test_invalid_invoice_json_leaves_no_outputs(tmp_path):
    invalid = tmp_path / "invalid.json"
    invalid.write_text("{not valid json", encoding="utf-8")
    output = tmp_path / "output"

    with pytest.raises(GenerationError, match="not valid JSON"):
        RmntcDocumentGenerator(ROOT).generate(invalid, output)

    assert not output.exists()


def test_missing_template_leaves_no_outputs(tmp_path):
    generator = RmntcDocumentGenerator(ROOT)
    generator.templates = (
        replace(
            generator.templates[0],
            template_path=tmp_path / "missing-statement.xlsx",
        ),
        *generator.templates[1:],
    )
    output = tmp_path / "output"

    with pytest.raises(GenerationError, match="Missing statement template"):
        generator.generate(SOURCE, output)

    assert not output.exists()


@pytest.mark.parametrize(
    ("item_count", "failed_documents"),
    (
        (7, {"comparison"}),
        (11, {"statement", "comparison"}),
        (12, {"statement", "quotation", "comparison"}),
    ),
)
def test_item_capacity_failures_leave_no_outputs(
    tmp_path, item_count, failed_documents
):
    source = _write_invoice(tmp_path / f"invoice-{item_count}.json", item_count)
    output = tmp_path / "output"

    with pytest.raises(GenerationError, match="template capacity exceeded") as caught:
        RmntcDocumentGenerator(ROOT).generate(source, output)

    assert {
        document
        for document, status in caught.value.statuses.items()
        if status == "failed"
    } == failed_documents
    assert not output.exists()


def test_renderer_failure_preserves_existing_outputs_and_leaves_no_partial_set(
    tmp_path,
):
    output = tmp_path / "output"
    output.mkdir()
    originals = {
        name: f"existing-{name}".encode()
        for name in ("statement.xlsx", "quotation.xlsx", "comparison.xlsx")
    }
    for name, payload in originals.items():
        (output / name).write_bytes(payload)

    generator = RmntcDocumentGenerator(ROOT)

    def fail_renderer(template, destination, model):
        raise RuntimeError("injected quotation failure")

    generator.renderers["quotation"] = fail_renderer

    with pytest.raises(GenerationError, match="injected quotation failure") as caught:
        generator.generate(SOURCE, output)

    assert caught.value.statuses == {
        "statement": "validated",
        "quotation": "failed",
        "comparison": "not run",
    }
    assert {
        name: (output / name).read_bytes() for name in originals
    } == originals


def test_comparison_uses_configured_markup(tmp_path):
    generator = RmntcDocumentGenerator(ROOT)
    report = generator.generate(SOURCE, tmp_path / "output")
    comparison = next(
        result for result in report.documents if result.document == "comparison"
    )

    assert report.comparison_markup == generator.configuration()[0]
    assert report.comparison_markup == Decimal("0.08")
    assert comparison.total == Decimal("3500000") * (
        Decimal(1) + report.comparison_markup
    )


def test_source_templates_remain_unchanged_after_generation(tmp_path):
    generator = RmntcDocumentGenerator(ROOT)
    before = _template_hashes(generator)

    generator.generate(SOURCE, tmp_path / "output")

    assert _template_hashes(generator) == before


def test_cli_reports_configuration_status_outputs_and_totals(tmp_path):
    output = tmp_path / "output"
    completed = subprocess.run(
        [
            sys.executable,
            "-m",
            "document_generator.cli",
            str(SOURCE),
            "--company",
            "rmntc",
            "--output-dir",
            str(output),
        ],
        cwd=ROOT,
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    for expected in (
        "Source JSON:",
        "Company: rmntc",
        "Template (statement):",
        "Template (quotation):",
        "Template (comparison):",
        "Comparison markup: 0.08 (8.00%)",
        "Statement: success",
        "Quotation: success",
        "Comparison: success",
        "Total: 3850000",
        "Total: 3780000",
        "Generation succeeded:",
    ):
        assert expected in completed.stdout
