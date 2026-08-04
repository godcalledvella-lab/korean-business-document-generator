from __future__ import annotations

import hashlib
import os
import re
import subprocess
import zipfile
from copy import deepcopy
from pathlib import Path
from xml.etree import ElementTree

import pytest
from openpyxl import load_workbook

from business import BusinessRuleEngine
from excel_renderer.comparison_writer import (
    COMPARISON_SHEET_PART,
    ITEM_ROWS as COMPARISON_ITEM_ROWS,
    write_comparison,
)
from excel_renderer.quotation_writer import (
    ITEM_ROWS as QUOTATION_ITEM_ROWS,
    QUOTATION_SHEET_PART,
    _wrap_template_text,
    korean_amount_words,
    write_quotation,
)


ROOT = Path(__file__).resolve().parents[1]
QUOTATION_TEMPLATE = ROOT / "reference/rmntc/templates/quotation.xlsx"
COMPARISON_TEMPLATE = ROOT / "reference/rmntc/templates/comparison.xlsx"
NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


@pytest.fixture()
def engine() -> BusinessRuleEngine:
    return BusinessRuleEngine(
        ROOT / "configs/invoice.schema.json",
        ROOT / "configs/business_rules.json",
    )


@pytest.fixture()
def invoice(engine):
    return engine.load_invoice(ROOT / "input/invoice.json")


def _parts(path: Path) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as archive:
        assert archive.testzip() is None
        return {name: archive.read(name) for name in archive.namelist()}


def _cell_pattern(reference: str) -> re.Pattern[bytes]:
    ref = re.escape(reference.encode())
    return re.compile(
        rb'<c\s[^>]*\br="' + ref + rb'"[^>]*?(?:/>|>.*?</c>)',
        re.DOTALL,
    )


def _without_cells(xml: bytes, references: set[str]) -> bytes:
    for reference in references:
        xml, count = _cell_pattern(reference).subn(
            b'<dynamic-cell r="' + reference.encode() + b'"/>', xml
        )
        assert count == 1, reference
    return xml


def _formula_map(xml: bytes) -> dict[str, tuple[str, tuple[tuple[str, str], ...]]]:
    root = ElementTree.fromstring(xml)
    formulas = {}
    for cell in root.findall(".//x:c", NS):
        formula = cell.find("x:f", NS)
        if formula is not None:
            formulas[cell.attrib["r"]] = (
                formula.text or "",
                tuple(sorted(formula.attrib.items())),
            )
    return formulas


def _merge_refs(xml: bytes) -> tuple[str, ...]:
    root = ElementTree.fromstring(xml)
    return tuple(
        cell.attrib["ref"] for cell in root.findall(".//x:mergeCell", NS)
    )


def _style_map(xml: bytes) -> dict[str, str | None]:
    root = ElementTree.fromstring(xml)
    return {
        cell.attrib["r"]: cell.attrib.get("s")
        for cell in root.findall(".//x:c", NS)
    }


def _structural_fragments(xml: bytes) -> dict[bytes, list[bytes]]:
    tags = (
        b"cols",
        b"mergeCells",
        b"pageMargins",
        b"pageSetup",
        b"printOptions",
        b"sheetPr",
        b"sheetViews",
    )
    result = {}
    for tag in tags:
        pattern = re.compile(
            rb"<" + tag + rb"(?:\s[^>]*)?(?:/>|>.*?</" + tag + rb">)",
            re.DOTALL,
        )
        result[tag] = pattern.findall(xml)
    result[b"row"] = re.findall(rb"<row(?:\s[^>]*)?>", xml)
    return result


def _preserved_parts(parts: dict[str, bytes]) -> set[str]:
    return {
        name
        for name in parts
        if name.startswith(
            (
                "xl/media/",
                "xl/drawings/",
                "xl/printerSettings/",
            )
        )
        or name
        in {
            "xl/styles.xml",
            "xl/theme/theme1.xml",
            "xl/workbook.xml",
            "xl/_rels/workbook.xml.rels",
            "docProps/core.xml",
            "docProps/custom.xml",
        }
        or name.startswith("xl/worksheets/_rels/")
    }


def _assert_only_mapped_cells_changed(
    source: Path,
    output: Path,
    sheet_part: str,
    modified_cells: tuple[str, ...],
    changed_formula_cells: tuple[str, ...] = (),
) -> None:
    before = _parts(source)
    after = _parts(output)
    assert before.keys() == after.keys()
    for name in before:
        if name != sheet_part:
            assert after[name] == before[name], name
    allowed = set(modified_cells)
    assert _without_cells(before[sheet_part], allowed) == _without_cells(
        after[sheet_part], allowed
    )
    assert _style_map(before[sheet_part]) == _style_map(after[sheet_part])
    assert _merge_refs(before[sheet_part]) == _merge_refs(after[sheet_part])
    assert _structural_fragments(before[sheet_part]) == _structural_fragments(
        after[sheet_part]
    )
    before_formulas = _formula_map(before[sheet_part])
    after_formulas = _formula_map(after[sheet_part])
    for reference in changed_formula_cells:
        before_formulas.pop(reference, None)
        after_formulas.pop(reference, None)
    assert before_formulas == after_formulas
    preserved = _preserved_parts(before)
    assert "xl/styles.xml" in preserved
    assert all(before[name] == after[name] for name in preserved)
    before_media = {
        name: hashlib.sha256(payload).hexdigest()
        for name, payload in before.items()
        if name.startswith("xl/media/")
    }
    after_media = {
        name: hashlib.sha256(payload).hexdigest()
        for name, payload in after.items()
        if name.startswith("xl/media/")
    }
    assert before_media == after_media


def test_quotation_values_totals_and_ooxml_are_preserved(tmp_path, engine, invoice):
    output = tmp_path / "quotation.xlsx"
    source_hash = hashlib.sha256(QUOTATION_TEMPLATE.read_bytes()).hexdigest()
    model = engine.create_quotation(invoice)
    report = write_quotation(QUOTATION_TEMPLATE, output, model)

    assert hashlib.sha256(QUOTATION_TEMPLATE.read_bytes()).hexdigest() == source_hash
    _assert_only_mapped_cells_changed(
        QUOTATION_TEMPLATE,
        output,
        QUOTATION_SHEET_PART,
        report.modified_cells,
    )

    workbook = load_workbook(output, data_only=False)
    sheet = workbook["견적서"]
    assert sheet["H2"].value.date().isoformat() == "2026-07-29"
    assert sheet["H3"].value == "주식회사 한빛상사\u00a0귀하"
    product_names = ", ".join(
        item["description"] for item in invoice["document"]["items"]
    )
    assert sheet["H4"].value == "업무 프로세스 분석 외 2건"
    assert sheet["B8"].value == product_names
    assert sheet["A6"].value.startswith("■ 상호 : 알엠엔티씨 주식회사")
    assert sheet["A6"].value.count("■") == 6
    assert "\n■ 주소" in sheet["A6"].value
    assert "로맨틱어스" not in sheet["A6"].value
    assert invoice["document"]["seller"]["name"] in sheet["A6"].value
    assert (
        invoice["document"]["seller"]["business_registration_number"]
        in sheet["A6"].value
    )
    assert invoice["document"]["seller"]["business_type"] in sheet["A6"].value
    assert invoice["document"]["seller"]["business_item"] in sheet["A6"].value
    assert sheet["C14"].value == "삼백오십만 원 정"
    assert sheet["F14"].value == "=F27"
    assert sheet["F27"].value == "=SUM(F16:G26)"
    assert sheet["F28"].value == 350000
    assert sheet["F29"].value == "=F27+F28"

    items = invoice["document"]["items"]
    for row, item in zip(QUOTATION_ITEM_ROWS, items):
        assert sheet[f"A{row}"].value == item["line_number"]
        assert sheet[f"B{row}"].value == item["description"]
        assert sheet[f"D{row}"].value == item["quantity"]
        assert sheet[f"E{row}"].value == item["unit_price"]
        assert sheet[f"H{row}"].value == item["remarks"]
    assert sum(item["supply_amount"] for item in items) == 3500000
    assert invoice["document"]["totals"]["total"] == 3850000

    for row in QUOTATION_ITEM_ROWS[len(items) :]:
        for column in ("A", "B", "D", "E", "H"):
            assert sheet[f"{column}{row}"].value is None
            assert sheet[f"{column}{row}"].style_id != 0
    workbook.close()
    calculated = load_workbook(output, data_only=True)["견적서"]
    assert [calculated[f"F{row}"].value for row in (16, 17, 18)] == [
        1500000,
        1500000,
        500000,
    ]
    assert calculated["F14"].value == 3500000
    assert calculated["F27"].value == 3500000
    assert calculated["F29"].value == 3850000


def test_long_quotation_client_uses_controlled_line_break():
    value = "주식회사 아주긴한글회사이름 테스트사업부 귀하"
    wrapped = _wrap_template_text(value)

    assert "\n" in wrapped
    assert wrapped.replace("\n", " ") == value


def test_comparison_markup_totals_and_ooxml_are_preserved(
    tmp_path, engine, invoice
):
    output = tmp_path / "comparison.xlsx"
    source_hash = hashlib.sha256(COMPARISON_TEMPLATE.read_bytes()).hexdigest()
    model = engine.create_comparison(invoice)
    report = write_comparison(COMPARISON_TEMPLATE, output, model)

    assert hashlib.sha256(COMPARISON_TEMPLATE.read_bytes()).hexdigest() == source_hash
    _assert_only_mapped_cells_changed(
        COMPARISON_TEMPLATE,
        output,
        COMPARISON_SHEET_PART,
        report.modified_cells,
        ("G20",),
    )

    workbook = load_workbook(output, data_only=False)
    sheet = workbook["Sheet1"]
    document = model.data["document"]
    rate = document["markup"]["rate"]
    assert rate == engine.pricing.comparison_markup_rate
    assert sheet["B6"].value == "2026.07.29"
    assert sheet["B7"].value == "주식회사 한빛상사"
    assert sheet["B11"].value == "=(G20)"
    assert sheet["G14"].value == "=C14*E14"
    assert sheet["G20"].value == "=SUM(G14:G19)"

    for row, source_item, comparison_item in zip(
        COMPARISON_ITEM_ROWS,
        invoice["document"]["items"],
        document["items"],
    ):
        assert sheet[f"A{row}"].value == source_item["description"]
        assert sheet[f"C{row}"].value == source_item["quantity"]
        assert sheet[f"D{row}"].value == source_item["unit"]
        assert sheet[f"E{row}"].value == comparison_item["comparison"]["unit_price"]
        if row != 14:
            assert (
                sheet[f"G{row}"].value
                == comparison_item["comparison"]["supply_amount"]
            )
        assert sheet[f"I{row}"].value == source_item["remarks"]

    expected_total = document["totals"]["comparison"]["supply_amount"]
    assert expected_total == 4158000

    for row in COMPARISON_ITEM_ROWS[len(document["items"]) :]:
        for column in ("A", "C", "D", "E", "G", "I"):
            assert sheet[f"{column}{row}"].value is None
            assert sheet[f"{column}{row}"].style_id != 0
    workbook.close()
    calculated = load_workbook(output, data_only=True)["Sheet1"]
    assert calculated["G14"].value == 1782000
    assert calculated["G20"].value == 4158000
    assert calculated["B11"].value == 4158000


def test_single_item_comparison_displays_marked_up_row_and_reconciled_total(
    tmp_path, engine, invoice
):
    reviewed = deepcopy(invoice)
    reviewed["document"]["items"] = [{
        "line_number": 1,
        "description": "트로피",
        "quantity": 8,
        "unit": "ea",
        "unit_price": 300000,
        "supply_amount": 2400000,
        "vat": 0,
        "total": 2400000,
        "remarks": "",
    }]
    reviewed["document"]["totals"] = {
        "supply_amount": 2400000,
        "vat": 0,
        "total": 2400000,
    }
    reviewed["extensions"] = {"rmntc.comparison_markup_percentage": 10}
    output = tmp_path / "single-comparison.xlsx"

    model = engine.create_comparison(reviewed)
    write_comparison(COMPARISON_TEMPLATE, output, model)
    sheet = load_workbook(output, data_only=True)["Sheet1"]

    assert sheet["E14"].value == 330000
    assert sheet["G14"].value == 2640000
    assert sheet["G20"].value == 2640000


def test_korean_amount_words():
    assert korean_amount_words(0) == "영"
    assert korean_amount_words(3_500_000) == "삼백오십만"
    assert korean_amount_words(123_456_789) == "일억이천삼백사십오만육천칠백팔십구"
    assert korean_amount_words(1_456_000) == "백사십오만육천"


@pytest.mark.skipif(
    os.environ.get("RMNTC_EXCEL_INTEGRATION") != "1",
    reason="set RMNTC_EXCEL_INTEGRATION=1 to open generated files in Microsoft Excel",
)
def test_generated_workbooks_open_successfully_in_microsoft_excel(
    tmp_path, engine, invoice
):
    quotation = tmp_path / "quotation.xlsx"
    comparison = tmp_path / "comparison.xlsx"
    write_quotation(
        QUOTATION_TEMPLATE, quotation, engine.create_quotation(invoice)
    )
    write_comparison(
        COMPARISON_TEMPLATE, comparison, engine.create_comparison(invoice)
    )
    script = """
on run argv
    tell application "Microsoft Excel"
        repeat with workbookPath in argv
            set openedWorkbook to open workbook workbook file name workbookPath
            close openedWorkbook saving no
        end repeat
    end tell
end run
""".strip()
    completed = subprocess.run(
        ["/usr/bin/osascript", "-e", script, str(quotation), str(comparison)],
        check=False,
        capture_output=True,
        text=True,
        timeout=120,
    )
    assert completed.returncode == 0, completed.stderr
