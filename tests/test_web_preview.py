from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]


def test_generation_adapter_creates_six_fallback_previews(tmp_path):
    session = tmp_path / "0123456789abcdef0123456789abcdef"
    session.mkdir()
    shutil.copy2(ROOT / "input/tax_invoice.pdf", session / "tax-invoice.pdf")
    stale = session / "generated"
    stale.mkdir()
    (stale / "final_package.pdf").write_bytes(b"stale package")
    (stale / "page-2-statement.pdf").write_bytes(b"stale preview")
    (session / "session.json").write_text(
        json.dumps(
            {
                "sessionId": session.name,
                "sourceName": "tax_invoice.pdf",
                "sourceType": "pdf",
                "draft": {},
                "confidences": {},
                "validation": {
                    "safeToApprove": True,
                    "schemaConformant": True,
                    "missing": [],
                    "lowConfidence": [],
                    "arithmeticMismatches": [],
                },
            }
        ),
        encoding="utf-8",
    )
    draft = json.loads((ROOT / "input/invoice.json").read_text(encoding="utf-8"))
    draft["extensions"] = {
        "rmntc.comparison_markup_percentage": 7.5,
        "rmntc.review_settings": {
            "statement": {
                "sender": "검토 발신자",
                "companyName": "검토 상호",
                "bank": "검토은행",
                "accountNumber": "123-456",
            },
            "blueQuotation": {
                "client": "사단법인 대한공도협회 귀하",
                "product": "트로피 외 2건",
                "showProductDetails": False,
                "productDetails": "트로피, 현수막, 모자",
                "showRemark": False,
                "remark": "VAT 별도",
                "informationBar": {
                    "companyName": {"value": "검토 공급자", "visible": True},
                    "businessNumber": {"value": "123-45-67890", "visible": True},
                    "address": {"value": "서울특별시", "visible": True},
                    "representative": {"value": "김대표", "visible": True},
                },
                "footer": {
                    "telephone": "02-1111-2222",
                    "email": "review@example.com",
                    "bank": "검토은행",
                    "accountNumber": "123-456",
                },
            },
            "greenQuotation": {
                "buyerCompany": "검토 구매자",
                "quotationDate": "2026-08-01",
                "companyProfile": {
                    "registrationNumber": "214-89-07571",
                    "companyName": "우현코퍼레이션",
                    "representative": "한주호",
                    "address": "부산광역시",
                    "businessType": "제조업",
                    "businessItem": "OEM ODM 제조",
                    "phone": "010-4480-7709",
                    "hp": "010-4480-7709",
                },
            },
        },
    }
    environment = {
        **os.environ,
        "PYTHONPATH": str(ROOT),
        "PDF_BACKEND": "libreoffice",
    }

    completed = subprocess.run(
        [
            sys.executable,
            str(ROOT / "web/scripts/generate_documents.py"),
            str(session),
            json.dumps(draft, ensure_ascii=False),
        ],
        cwd=ROOT,
        env=environment,
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    result = json.loads(completed.stdout)
    assert result["previewMode"] == "html"
    assert result["comparisonMarkupPercentage"] == 7.5
    assert set(result["previews"]) == {
        "taxInvoice",
        "statement",
        "quotation",
        "comparison",
        "businessRegistration",
        "bankAccount",
    }
    assert "package" not in result["downloads"]
    assert not (session / "generated" / "final_package.pdf").exists()
    assert not (session / "generated" / "page-2-statement.pdf").exists()
    assert result["downloads"]["bundle"].endswith("/bundle")
    bundle = session / "generated" / result["bundleName"]
    assert bundle.is_file()
    with zipfile.ZipFile(bundle) as archive:
        assert set(archive.namelist()) == {
            "01-tax-invoice.pdf",
            "approved-invoice.json",
            "02-statement.xlsx",
            "03-quotation.xlsx",
            "04-comparison-quotation.xlsx",
            "05-business-registration.pdf",
            "06-bank-account.pdf",
            "previews/02-statement.html",
            "previews/03-quotation.html",
            "previews/04-comparison-quotation.html",
        }
        approved = json.loads(archive.read("approved-invoice.json"))
        assert (
            approved["extensions"]["rmntc.comparison_markup_percentage"]
            == 7.5
        )
        assert (
            approved["extensions"]["rmntc.review_settings"]["blueQuotation"][
                "client"
            ]
            == "사단법인 대한공도협회 귀하"
        )
        assert archive.read("04-comparison-quotation.xlsx") == (
            session / "generated/comparison.xlsx"
        ).read_bytes()
    assert result["packageError"]
    for name in ("statement", "quotation", "comparison"):
        preview = session / "generated" / f"{name}.html"
        text = preview.read_text(encoding="utf-8")
        assert "<table>" in text
        assert 'class="page"' in text
    for name in ("business_registration.pdf", "bank_account.pdf"):
        assert len(PdfReader(session / "generated" / name).pages) == 1
    stored = json.loads((session / "session.json").read_text(encoding="utf-8"))
    assert stored["comparisonMarkupPercentage"] == 7.5
    comparison = load_workbook(
        session / "generated/comparison.xlsx",
        data_only=False,
    )
    assert comparison["Sheet1"]["E14"].value == 1_773_750
    assert comparison["Sheet1"]["B7"].value == "검토 구매자"
    assert comparison["Sheet1"]["F6"].value == "우현코퍼레이션"
    comparison.close()
    statement = load_workbook(
        session / "generated/statement.xlsx",
        data_only=False,
    )
    assert statement["청구서"]["B4"].value == "발신자: 검토 발신자"
    assert statement["청구서"]["C20"].value == "검토 상호"
    assert statement["청구서"]["C21"].value == "검토은행"
    assert statement["청구서"]["C22"].value == "123-456"
    statement.close()
    quotation = load_workbook(
        session / "generated/quotation.xlsx",
        data_only=False,
    )
    assert quotation["견적서"]["H3"].value.replace("\n", " ").replace(
        "\u00a0", " "
    ) == (
        "사단법인 대한공도협회 귀하"
    )
    assert quotation["견적서"]["B8"].value is None
    assert quotation["견적서"]["B30"].value is None
    assert quotation["견적서"]["A35"].value.endswith(
        "계좌번호 : 검토은행 123-456"
    )
    quotation.close()
    assert stored["reviewSettings"] == draft["extensions"]["rmntc.review_settings"]


def test_workbook_preview_does_not_change_source_hash(tmp_path):
    import hashlib
    import importlib.util

    module_path = ROOT / "web/scripts/workbook_preview.py"
    spec = importlib.util.spec_from_file_location("workbook_preview", module_path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    source = ROOT / "output/quotation.xlsx"
    before = hashlib.sha256(source.read_bytes()).hexdigest()

    module.render_workbook_preview(
        source,
        tmp_path / "quotation.html",
        title="Quotation",
    )

    assert hashlib.sha256(source.read_bytes()).hexdigest() == before
    preview = (tmp_path / "quotation.html").read_text(encoding="utf-8")
    assert "font-family:'Malgun Gothic'" in preview


def test_statement_preview_materializes_template_stripes_and_centers_sheet(tmp_path):
    import importlib.util

    module_path = ROOT / "web/scripts/workbook_preview.py"
    spec = importlib.util.spec_from_file_location("workbook_preview", module_path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    output = tmp_path / "statement.html"
    module.render_workbook_preview(
        ROOT / "output/statement.xlsx",
        output,
        title="Statement",
    )

    preview = output.read_text(encoding="utf-8")
    assert "background:#F8F5EE" in preview
    assert "margin:0 auto" in preview
