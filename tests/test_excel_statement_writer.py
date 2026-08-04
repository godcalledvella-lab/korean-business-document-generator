from __future__ import annotations

import hashlib
import re
import unicodedata
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import pytest
from openpyxl import load_workbook

from business import BusinessRuleEngine, ViewModel, ViewModelType
from excel_renderer import ItemCapacityError, write_statement
from excel_renderer.statement_writer import STATEMENT_SHEET_PART


ROOT = Path(__file__).resolve().parents[1]
NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _statement_template() -> Path:
    matches = [
        path
        for path in (ROOT / "reference/rmntc/templates").glob("*.xlsx.xlsx")
        if "거래명세서" in unicodedata.normalize("NFC", path.name)
    ]
    assert len(matches) == 1
    return matches[0]


@pytest.fixture()
def statement_view_model() -> ViewModel:
    engine = BusinessRuleEngine(
        ROOT / "configs/invoice.schema.json",
        ROOT / "configs/business_rules.json",
    )
    return engine.create_statement(engine.load_invoice(ROOT / "input/invoice.json"))


def _parts(path: Path) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as archive:
        assert archive.testzip() is None
        return {name: archive.read(name) for name in archive.namelist()}


def _cell_pattern(reference: str) -> re.Pattern[bytes]:
    ref = re.escape(reference.encode())
    return re.compile(
        rb'<c\s[^>]*\br="' + ref + rb'"[^>]*?(?:/>|>.*?</c>)',
        re.DOTALL,
    )


def _without_cells(xml: bytes, references: set[str]) -> bytes:
    for reference in references:
        xml, count = _cell_pattern(reference).subn(
            b'<dynamic-cell r="' + reference.encode() + b'"/>', xml
        )
        assert count == 1, reference
    return xml


def _formula_map(xml: bytes) -> dict[str, str]:
    root = ElementTree.fromstring(xml)
    formulas: dict[str, str] = {}
    for cell in root.findall(".//x:c", NS):
        formula = cell.find("x:f", NS)
        if formula is not None:
            formulas[cell.attrib["r"]] = formula.text or ""
    return formulas


def _merge_refs(xml: bytes) -> tuple[str, ...]:
    root = ElementTree.fromstring(xml)
    return tuple(
        cell.attrib["ref"] for cell in root.findall(".//x:mergeCell", NS)
    )


def test_only_mapped_statement_cells_change(tmp_path, statement_view_model):
    source = _statement_template()
    output = tmp_path / "statement.xlsx"
    source_hash = hashlib.sha256(source.read_bytes()).hexdigest()

    report = write_statement(source, output, statement_view_model)
    before = _parts(source)
    after = _parts(output)

    assert hashlib.sha256(source.read_bytes()).hexdigest() == source_hash
    assert before.keys() == after.keys()
    for name in before:
        if name != STATEMENT_SHEET_PART:
            assert after[name] == before[name], name

    allowed = set(report.modified_cells)
    assert _without_cells(before[STATEMENT_SHEET_PART], allowed) == _without_cells(
        after[STATEMENT_SHEET_PART], allowed
    )
    assert set(report.created_formula_cells) == {"E8", "E10"}


def test_merges_images_styles_dimensions_and_settings_are_unchanged(
    tmp_path, statement_view_model
):
    source = _statement_template()
    output = tmp_path / "statement.xlsx"
    write_statement(source, output, statement_view_model)
    before = _parts(source)
    after = _parts(output)

    assert _merge_refs(before[STATEMENT_SHEET_PART]) == _merge_refs(
        after[STATEMENT_SHEET_PART]
    )
    preserved_parts = {
        name
        for name in before
        if name.startswith(("xl/media/", "xl/drawings/"))
        or name in {
            "xl/styles.xml",
            "xl/workbook.xml",
            "xl/worksheets/sheet2.xml",
        }
    }
    assert preserved_parts
    assert all(before[name] == after[name] for name in preserved_parts)

    for tag in (b"cols", b"pageMargins", b"pageSetup", b"printOptions"):
        pattern = re.compile(rb"<" + tag + rb"(?:\s[^>]*)?(?:/>|>.*?</" + tag + rb">)")
        assert pattern.findall(before[STATEMENT_SHEET_PART]) == pattern.findall(
            after[STATEMENT_SHEET_PART]
        )


def test_existing_formulas_are_untouched_and_only_blank_item_formulas_are_added(
    tmp_path, statement_view_model
):
    source = _statement_template()
    output = tmp_path / "statement.xlsx"
    write_statement(source, output, statement_view_model)
    before = _parts(source)
    after = _parts(output)

    original = _formula_map(before[STATEMENT_SHEET_PART])
    rendered = _formula_map(after[STATEMENT_SHEET_PART])
    assert rendered["E7"] == original["E7"]
    assert "D2" not in rendered
    assert "E18" not in rendered
    assert rendered["E8"] == "(D8*B8)"
    assert rendered["E10"] == "(D10*B10)"
    assert set(rendered) - set(original) == {"E8", "E10"}


def test_generated_workbook_is_a_valid_openable_ooxml_package(
    tmp_path, statement_view_model
):
    output = tmp_path / "statement.xlsx"
    write_statement(_statement_template(), output, statement_view_model)

    assert zipfile.is_zipfile(output)
    parts = _parts(output)
    for required in (
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        STATEMENT_SHEET_PART,
    ):
        ElementTree.fromstring(parts[required])
    assert b'name="\xec\xb2\xad\xea\xb5\xac\xec\x84\x9c"' in parts["xl/workbook.xml"]

    # OpenPyXL is an independent, Excel-compatible OOXML parser. Loading the
    # result catches invalid relationships or worksheet records without saving
    # (and therefore without normalizing) the generated workbook.
    workbook = load_workbook(output, data_only=False)
    statement = workbook["청구서"]
    assert workbook.sheetnames == ["청구서", "Sheet1"]
    assert len(statement._images) == 1
    expected_total = statement_view_model.data["document"]["totals"]["total"]
    assert statement["D2"].value == expected_total
    assert statement["E7"].value == "=(D7*B7)"
    assert statement["E18"].value == expected_total
    workbook.close()


def test_capacity_error_does_not_create_output(tmp_path, statement_view_model):
    document = dict(statement_view_model.data)
    document["document"] = dict(document["document"])
    document["document"]["items"] = list(document["document"]["items"]) * 4
    oversized = ViewModel(ViewModelType.STATEMENT, document)
    output = tmp_path / "too-many.xlsx"

    with pytest.raises(ItemCapacityError, match="10 styled item rows"):
        write_statement(_statement_template(), output, oversized)
    assert not output.exists()
