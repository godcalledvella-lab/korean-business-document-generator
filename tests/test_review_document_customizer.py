from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from document_generator import RmntcDocumentGenerator
from excel_renderer.review_customizer import apply_review_settings


ROOT = Path(__file__).resolve().parents[1]


def _settings() -> dict:
    return {
        "statement": {
            "sender": "검토 발신자",
            "companyName": "검토 상호",
            "bank": "검토은행",
            "accountNumber": "123-456",
        },
        "blueQuotation": {
            "client": "사단법인 대한공도협회 귀하",
            "product": "트로피 외 5건",
            "showProductDetails": True,
            "productDetails": "트로피, 현수막, 모자, 책자, 우산, 배지",
            "showRemark": True,
            "remark": "VAT 별도",
            "informationBar": {
                "companyName": {"value": "송장 공급자", "visible": True},
                "businessNumber": {"value": "123-45-67890", "visible": True},
                "address": {"value": "서울특별시", "visible": True},
                "representative": {"value": "김대표", "visible": True},
            },
            "footer": {
                "telephone": "02-1111-2222",
                "email": "review@example.com",
                "bank": "검토은행",
                "accountNumber": "123-456",
            },
        },
        "greenQuotation": {
            "buyerCompany": "사단법인 대한공도협회",
            "quotationDate": "2026-07-30",
            "companyProfile": {
                "registrationNumber": "214-89-07571",
                "companyName": "우현코퍼레이션",
                "representative": "한주호",
                "address": "부산광역시",
                "businessType": "제조업",
                "businessItem": "OEM ODM 제조",
                "phone": "010-4480-7709",
                "hp": "010-4480-7709",
            },
        },
    }


def _parts(path: Path) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


def _structure(xml: bytes) -> tuple[list[bytes], list[bytes], list[bytes]]:
    return (
        re.findall(rb"<mergeCell[^>]*/>", xml),
        re.findall(rb"<row(?:\s[^>]*)?>", xml),
        re.findall(
            rb"<(?:cols|pageMargins|pageSetup|printOptions|drawing|tableParts)"
            rb"(?:\s[^>]*)?(?:/>|>.*?</(?:cols|pageMargins|pageSetup|"
            rb"printOptions|drawing|tableParts)>)",
            xml,
            re.DOTALL,
        ),
    )


def test_review_settings_change_only_mapped_values(tmp_path):
    generated = tmp_path / "generated"
    RmntcDocumentGenerator(ROOT).generate(ROOT / "input/invoice.json", generated)
    before_statement = _parts(generated / "statement.xlsx")
    before_quotation = _parts(generated / "quotation.xlsx")
    before_comparison = _parts(generated / "comparison.xlsx")

    apply_review_settings(
        generated / "statement.xlsx",
        generated / "quotation.xlsx",
        generated / "comparison.xlsx",
        _settings(),
    )

    statement = load_workbook(generated / "statement.xlsx", data_only=False)
    sheet = statement["청구서"]
    assert sheet["B4"].value == "발신자: 검토 발신자"
    assert sheet["C20"].value == "검토 상호"
    assert sheet["C21"].value == "검토은행"
    assert sheet["C22"].value == "123-456"
    statement.close()

    quotation = load_workbook(generated / "quotation.xlsx", data_only=False)
    blue = quotation["견적서"]
    assert blue["H3"].value.replace("\n", " ").replace("\u00a0", " ") == (
        "사단법인 대한공도협회 귀하"
    )
    assert blue["H4"].value == "트로피 외 5건"
    assert blue["B8"].value == "트로피, 현수막, 모자, 책자, 우산, 배지"
    assert blue["B30"].value == "VAT 별도"
    assert "■ 업종 : 서비스업" in blue["A6"].value
    assert "■ 업태 : 경영 컨설팅 및 소프트웨어 개발" in blue["A6"].value
    assert "\n■ 주소 : 서울특별시" in blue["A6"].value
    assert blue["A6"].value.count("■") == 6
    assert blue["A35"].value == (
        "TEL : 02-1111-2222 / E-mail : review@example.com"
        " / 계좌번호 : 검토은행 123-456"
    )
    quotation.close()

    comparison = load_workbook(generated / "comparison.xlsx", data_only=False)
    green = comparison["Sheet1"]
    assert green["B6"].value == "2026.07.30"
    assert green["B7"].value == "사단법인 대한공도협회"
    assert green["F5"].value == "214-89-07571"
    assert green["F6"].value == "우현코퍼레이션"
    assert green["I6"].value == "한주호"
    assert green["F7"].value == "부산광역시"
    assert green["F8"].value == "제조업"
    assert green["I8"].value == "OEM ODM 제조"
    assert green["F9"].value == green["I9"].value == "010-4480-7709"
    assert green["B11"].value == "사백십오만팔천"
    assert green["G20"].value == "=SUM(G14:G19)"
    comparison.close()

    after_statement = _parts(generated / "statement.xlsx")
    after_quotation = _parts(generated / "quotation.xlsx")
    after_comparison = _parts(generated / "comparison.xlsx")
    for before, after, sheet_part in (
        (before_statement, after_statement, "xl/worksheets/sheet1.xml"),
        (before_quotation, after_quotation, "xl/worksheets/sheet1.xml"),
        (before_comparison, after_comparison, "xl/worksheets/sheet1.xml"),
    ):
        assert before.keys() == after.keys()
        assert all(
            before[name] == after[name]
            for name in before
            if name != sheet_part
        )
        assert _structure(before[sheet_part]) == _structure(after[sheet_part])


def test_review_defaults_use_fixed_rmntc_display_values():
    source = (
        "import fs from 'node:fs';"
        "import {defaultReviewDocumentSettings} from "
        f"{json.dumps((ROOT / 'web/lib/reviewDocuments.ts').as_uri())};"
        "const draft=JSON.parse(fs.readFileSync(0,'utf8'));"
        "const confidence={"
        "'document.seller.name':0.99,"
        "'document.seller.business_registration_number':0.99,"
        "'document.seller.representative':0.99,"
        "'document.seller.address':0.99,"
        "'document.seller.business_type':0.99,"
        "'document.seller.business_item':0.99,"
        "'document.seller.contact.phone':0.99,"
        "'document.seller.contact.email':0.99};"
        "process.stdout.write(JSON.stringify(defaultReviewDocumentSettings(draft,confidence)));"
    )
    import subprocess

    completed = subprocess.run(
        [
            "node",
            "--disable-warning=MODULE_TYPELESS_PACKAGE_JSON",
            "--experimental-strip-types",
            "--input-type=module",
            "-e",
            source,
        ],
        input=(ROOT / "input/invoice.json").read_text(encoding="utf-8"),
        text=True,
        capture_output=True,
        check=True,
    )
    settings = json.loads(completed.stdout)
    assert settings["statement"] == {
        "sender": "로맨틱어스",
        "companyName": "로맨틱어스",
        "bank": "신한은행",
        "accountNumber": "110-427-856988",
    }
    assert settings["blueQuotation"]["client"] == "주식회사 한빛상사 귀하"
    assert settings["blueQuotation"]["product"] == "업무 프로세스 분석 외 2건"
    assert settings["greenQuotation"]["buyerCompany"] == "주식회사 한빛상사"
    assert settings["greenQuotation"]["quotationDate"] == "2026-07-29"
    assert settings["greenQuotation"]["companyProfile"] == {
        "registrationNumber": "214-89-07571",
        "companyName": "우현코퍼레이션",
        "representative": "한주호",
        "address": "부산광역시",
        "businessType": "제조업",
        "businessItem": "OEM ODM 제조",
        "phone": "010-4480-7709",
        "hp": "010-4480-7709",
    }
    assert settings["blueQuotation"]["footer"] == {
        "telephone": "010-8579-0342",
        "email": "bigthumbdesigner@gmail.com",
        "bank": "신한은행",
        "accountNumber": "110-427-856988",
    }
    assert settings["blueQuotation"]["informationBar"] == {
        "companyName": {"value": "로맨틱어스", "visible": True},
        "businessNumber": {"value": "102-21-34572", "visible": True},
        "address": {
            "value": "경상남도 창원시 성산구 외동반림로126번길 57, 1층",
            "visible": True,
        },
        "representative": {"value": "정성우", "visible": True},
    }


def test_low_confidence_unknown_seller_keeps_template_defaults():
    source = (
        "import fs from 'node:fs';"
        "import {defaultReviewDocumentSettings} from "
        f"{json.dumps((ROOT / 'web/lib/reviewDocuments.ts').as_uri())};"
        "const draft=JSON.parse(fs.readFileSync(0,'utf8'));"
        "process.stdout.write(JSON.stringify(defaultReviewDocumentSettings("
        "draft,{'document.seller.name':0.2})));"
    )
    import subprocess

    draft = json.loads((ROOT / "input/invoice.json").read_text(encoding="utf-8"))
    draft["document"]["seller"]["name"] = "알 수 없는 공급자"
    completed = subprocess.run(
        [
            "node",
            "--disable-warning=MODULE_TYPELESS_PACKAGE_JSON",
            "--experimental-strip-types",
            "--input-type=module",
            "-e",
            source,
        ],
        input=json.dumps(draft, ensure_ascii=False),
        text=True,
        capture_output=True,
        check=True,
    )
    settings = json.loads(completed.stdout)
    assert settings["statement"] == {
        "sender": "로맨틱어스",
        "companyName": "로맨틱어스",
        "bank": "신한은행",
        "accountNumber": "110-427-856988",
    }
    assert settings["blueQuotation"]["informationBar"]["companyName"]["value"] == (
        "로맨틱어스"
    )
    assert settings["blueQuotation"]["footer"] == {
        "telephone": "010-8579-0342",
        "email": "bigthumbdesigner@gmail.com",
        "bank": "신한은행",
        "accountNumber": "110-427-856988",
    }
