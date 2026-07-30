from __future__ import annotations

import hashlib
import shutil
import subprocess
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from pypdf import PdfReader
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen.canvas import Canvas

from package.backends.base import PDFBackend
from package.backends.libreoffice import (
    LibreOfficeBackend,
    _fit_single_page_to_a4_portrait,
    _isolate_worksheet,
    _prepare_statement_render_copy,
    _target_worksheet,
    build_libreoffice_command,
    find_libreoffice,
)
from package.backends.microsoft_excel import MicrosoftExcelBackend
from package.backends.registry import BackendRegistry
from package.models import (
    BackendAvailability,
    PackageError,
    PackageInputs,
    RenderResult,
)
from package.service import PDFPackageGenerator


ROOT = Path(__file__).resolve().parents[1]


def _pdf(path: Path, *labels: str, page_size=A4) -> Path:
    canvas = Canvas(str(path), pagesize=page_size)
    for label in labels:
        canvas.drawString(72, 720, label)
        canvas.showPage()
    canvas.save()
    return path


def _inputs(tmp_path: Path) -> PackageInputs:
    workbooks = []
    for name in ("statement.xlsx", "quotation.xlsx", "comparison.xlsx"):
        path = tmp_path / name
        path.write_bytes(f"xlsx:{name}".encode())
        workbooks.append(path)
    return PackageInputs(
        _pdf(tmp_path / "tax.pdf", "tax"),
        *workbooks,
        _pdf(tmp_path / "business.pdf", "business"),
        _pdf(tmp_path / "bank.pdf", "bank"),
    )


class FakeBackend(PDFBackend):
    name = "fake"

    def __init__(
        self,
        *,
        available: bool = True,
        corrupt_for: str | None = None,
        pages_for: tuple[str, int] | None = None,
        modify_source: bool = False,
    ) -> None:
        self.available = available
        self.corrupt_for = corrupt_for
        self.pages_for = pages_for
        self.modify_source = modify_source

    def availability(self):
        return BackendAvailability(self.name, self.available, "fake backend")

    def render_xlsx(self, input_path, output_path):
        before = _hash(input_path)
        command = ("fake-render", str(input_path), str(output_path))
        if self.modify_source:
            input_path.write_bytes(input_path.read_bytes() + b"changed")
        if self.corrupt_for == input_path.stem:
            output_path.write_bytes(b"not a pdf")
        else:
            count = (
                self.pages_for[1]
                if self.pages_for and self.pages_for[0] == input_path.stem
                else 1
            )
            _pdf(output_path, *(f"{input_path.stem}-{i}" for i in range(count)))
        return RenderResult(
            self.name,
            input_path,
            output_path,
            command,
            0,
            "fake stdout",
            "",
            before,
            _hash(input_path),
        )


class NamedBackend(FakeBackend):
    def __init__(self, name, available):
        super().__init__(available=available)
        self.name = name


def _hash(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_backend_priority_and_explicit_selection():
    libreoffice = NamedBackend("libreoffice", True)
    excel = NamedBackend("excel", True)
    registry = BackendRegistry((excel, libreoffice))

    assert registry.select("auto")[0] is libreoffice
    assert registry.select("excel")[0] is excel


def test_auto_uses_excel_only_when_libreoffice_is_unavailable():
    libreoffice = NamedBackend("libreoffice", False)
    excel = NamedBackend("excel", True)
    assert BackendRegistry((libreoffice, excel)).select("auto")[0] is excel


def test_explicit_unavailable_and_unsupported_backend_fail_clearly():
    registry = BackendRegistry((NamedBackend("libreoffice", False),))
    with pytest.raises(PackageError, match="unavailable"):
        registry.select("libreoffice")
    with pytest.raises(PackageError, match="Unsupported"):
        registry.select("unknown")


def test_libreoffice_detection_command_and_isolated_profile(tmp_path):
    executable = tmp_path / "soffice"
    executable.write_text("#!/bin/sh\n", encoding="utf-8")
    executable.chmod(0o755)
    assert find_libreoffice((executable,), which=lambda _: None) == executable.resolve()

    command = build_libreoffice_command(
        executable,
        tmp_path / "copy.xlsx",
        tmp_path / "output",
        tmp_path / "isolated profile",
    )
    assert command[0] == str(executable)
    assert "--headless" in command
    assert "--convert-to" in command
    assert "pdf" in command
    profile = next(value for value in command if value.startswith("-env:"))
    assert "UserInstallation=file://" in profile
    assert "isolated%20profile" in profile


def test_libreoffice_detects_package_staging_names():
    assert _target_worksheet("statement") == "청구서"
    assert _target_worksheet("page-2-statement") == "청구서"
    assert _target_worksheet("page-3-quotation") == "견적서"
    assert _target_worksheet("page-4-comparison") == "Sheet1"


def test_libreoffice_backend_preserves_source_hash(tmp_path):
    executable = tmp_path / "soffice"
    executable.write_text("#!/bin/sh\n", encoding="utf-8")
    executable.chmod(0o755)
    source = tmp_path / "statement.xlsx"
    workbook = Workbook()
    workbook.active.title = "청구서"
    workbook.create_sheet("Sheet1")
    workbook.save(source)
    workbook.close()
    before = _hash(source)

    def runner(command, **kwargs):
        if "--version" in command:
            return subprocess.CompletedProcess(command, 0, "LibreOffice 25.2", "")
        output_dir = Path(command[command.index("--outdir") + 1])
        copied = Path(command[-1])
        _pdf(output_dir / copied.with_suffix(".pdf").name, "statement")
        return subprocess.CompletedProcess(command, 0, "converted", "")

    backend = LibreOfficeBackend(executable, runner=runner)
    result = backend.render_xlsx(source, tmp_path / "statement.pdf")

    assert result.source_hash_before == before == result.source_hash_after
    assert _hash(source) == before


def test_libreoffice_copy_isolates_registered_output_sheet(tmp_path):
    path = tmp_path / "statement.xlsx"
    workbook = Workbook()
    workbook.active.title = "청구서"
    workbook.create_sheet("Sheet1")
    workbook.save(path)
    workbook.close()

    _isolate_worksheet(path, "청구서")

    from openpyxl import load_workbook

    isolated = load_workbook(path)
    assert isolated["청구서"].sheet_state == "visible"
    assert isolated["Sheet1"].sheet_state == "hidden"
    isolated.close()


def test_libreoffice_copy_removes_only_pathological_duplicate_shapes(tmp_path):
    source = (
        ROOT / "reference/rmntc/templates/quotation.xlsx"
    )
    copied = tmp_path / "quotation.xlsx"
    shutil.copy2(source, copied)
    before = _hash(source)

    _isolate_worksheet(copied, "견적서")

    with zipfile.ZipFile(copied) as archive:
        drawing = archive.read("xl/drawings/drawing1.xml")
        assert drawing.count(b"<xdr:sp>") == 0
        assert drawing.count(b"<xdr:pic>") == 1
    assert _hash(source) == before


def test_libreoffice_statement_copy_materializes_template_stripes(tmp_path):
    source = ROOT / "reference/rmntc/templates/statement.xlsx"
    copied = tmp_path / "statement.xlsx"
    shutil.copy2(source, copied)
    before = _hash(source)

    _prepare_statement_render_copy(copied)

    from openpyxl import load_workbook

    workbook = load_workbook(copied)
    sheet = workbook["청구서"]
    for row in (7, 10, 12, 14, 16):
        assert sheet[f"B{row}"].fill.fgColor.rgb == "FFF8F5EE"
    for row in (8, 11, 13, 15, 17):
        assert sheet[f"B{row}"].fill.fgColor.rgb in {"00000000", "000000"}
    assert sheet["J17"].value is None
    assert str(sheet.print_area) == "'청구서'!$B$1:$E$24"
    assert sheet.page_setup.scale == 95
    assert sheet.page_setup.fitToWidth is None
    assert sheet.page_setup.fitToHeight is None
    workbook.close()
    assert _hash(source) == before


def test_libreoffice_landscape_page_is_vector_fitted_to_a4(tmp_path):
    path = _pdf(
        tmp_path / "comparison.pdf",
        "comparison-vector-content",
        page_size=(792, 612),
    )

    _fit_single_page_to_a4_portrait(path)

    page = PdfReader(path).pages[0]
    assert float(page.mediabox.width) == pytest.approx(595.2756, abs=0.1)
    assert float(page.mediabox.height) == pytest.approx(841.8898, abs=0.1)
    assert "comparison-vector-content" in (page.extract_text() or "")


def test_microsoft_excel_availability_failure(monkeypatch):
    monkeypatch.setattr(
        "package.backends.microsoft_excel.find_excel_application", lambda: None
    )
    availability = MicrosoftExcelBackend().availability()
    assert not availability.available
    assert "not found" in availability.reason


@pytest.mark.parametrize(
    ("mutation", "match"),
    (
        ("invalid_page_one", "invalid header"),
        ("missing_fixed", "PDF is missing"),
        ("corrupt_intermediate", "invalid header"),
        ("multipage_intermediate", "Expected 1 page"),
    ),
)
def test_validation_failures_leave_no_final_output(tmp_path, mutation, match):
    inputs = _inputs(tmp_path)
    backend = FakeBackend()
    if mutation == "invalid_page_one":
        inputs.tax_invoice.write_bytes(b"invalid")
    elif mutation == "missing_fixed":
        inputs.bank_account.unlink()
    elif mutation == "corrupt_intermediate":
        backend.corrupt_for = "quotation"
    elif mutation == "multipage_intermediate":
        backend.pages_for = ("comparison", 2)
    output = tmp_path / "final.pdf"

    with pytest.raises(PackageError, match=match):
        PDFPackageGenerator(
            BackendRegistry((backend,)), logger=lambda _: None
        ).generate(inputs, output, backend="fake")

    assert not output.exists()
    assert not list(tmp_path.glob(".pdf-package-*"))


def test_source_mutation_is_detected_and_existing_final_is_preserved(tmp_path):
    inputs = _inputs(tmp_path)
    output = tmp_path / "final.pdf"
    output.write_bytes(b"existing final")

    with pytest.raises(PackageError, match="Source XLSX changed") as caught:
        PDFPackageGenerator(
            BackendRegistry((FakeBackend(modify_source=True),)),
            logger=lambda _: None,
        ).generate(inputs, output, backend="fake")

    assert output.read_bytes() == b"existing final"
    assert caught.value.sources_unchanged is False


def test_successful_package_has_exact_order_six_pages_and_logs(tmp_path):
    inputs = _inputs(tmp_path)
    logs = []
    output = tmp_path / "final.pdf"
    report = PDFPackageGenerator(
        BackendRegistry((FakeBackend(),)), logger=logs.append
    ).generate(inputs, output, backend="fake")

    reader = PdfReader(output)
    assert report.page_count == 6
    assert len(reader.pages) == 6
    assert [page.extract_text().strip() for page in reader.pages] == [
        "tax",
        "statement-0",
        "quotation-0",
        "comparison-0",
        "business",
        "bank",
    ]
    assert logs[0] == "[1/8] Detecting PDF backend..."
    assert any("[8/8] Validating final package..." == line for line in logs)
    assert logs[-1] == f"✓ Saved to {output.resolve()}"
    assert not list(tmp_path.glob(".pdf-package-*"))


def test_detailed_error_diagnostics_include_execution_context(tmp_path):
    inputs = _inputs(tmp_path)
    with pytest.raises(PackageError) as caught:
        PDFPackageGenerator(
            BackendRegistry((FakeBackend(corrupt_for="statement"),)),
            logger=lambda _: None,
        ).generate(inputs, tmp_path / "final.pdf", backend="fake")

    diagnostics = caught.value.diagnostics()
    for expected in (
        "Failed step:",
        "Backend: fake",
        "Command:",
        "Exit code:",
        "Stdout:",
        "Stderr:",
        "Temporary working path:",
        "Source files unchanged: yes",
    ):
        assert expected in diagnostics
