"""Atomic orchestration for the three RMNTC Excel documents."""

from __future__ import annotations

import hashlib
import os
import re
import tempfile
import zipfile
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Callable, Mapping
from xml.etree import ElementTree

from openpyxl import load_workbook

from business import BusinessRuleEngine, ViewModel
from excel_renderer.comparison_writer import write_comparison
from excel_renderer.quotation_writer import write_quotation
from excel_renderer.statement_writer import write_statement


WORKSHEET_NS = {
    "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
}


class GenerationError(RuntimeError):
    """Raised when an atomic generation run cannot be completed."""

    def __init__(
        self,
        message: str,
        *,
        statuses: Mapping[str, str] | None = None,
    ) -> None:
        super().__init__(message)
        self.statuses = dict(statuses or {})


@dataclass(frozen=True)
class TemplateSpec:
    document: str
    template_path: Path
    output_name: str
    worksheet_name: str
    sheet_part: str


@dataclass(frozen=True)
class DocumentResult:
    document: str
    template_path: Path
    output_path: Path
    total: Decimal
    status: str


@dataclass(frozen=True)
class GenerationReport:
    source_path: Path
    company: str
    comparison_markup: Decimal
    documents: tuple[DocumentResult, ...]


class RmntcDocumentGenerator:
    """Generate, validate, and atomically publish one RMNTC workbook set."""

    company = "rmntc"

    def __init__(
        self,
        project_root: str | Path | None = None,
        *,
        renderers: Mapping[str, Callable[[Path, Path, ViewModel], object]]
        | None = None,
    ) -> None:
        self.project_root = (
            Path(project_root).resolve()
            if project_root is not None
            else Path(__file__).resolve().parents[1]
        )
        self.schema_path = self.project_root / "configs/invoice.schema.json"
        self.business_rules_path = self.project_root / "configs/business_rules.json"
        template_root = self.project_root / "reference/rmntc/templates"
        self.templates = (
            TemplateSpec(
                "statement",
                template_root / "statement.xlsx",
                "statement.xlsx",
                "청구서",
                "xl/worksheets/sheet1.xml",
            ),
            TemplateSpec(
                "quotation",
                template_root / "quotation.xlsx",
                "quotation.xlsx",
                "견적서",
                "xl/worksheets/sheet1.xml",
            ),
            TemplateSpec(
                "comparison",
                template_root / "comparison.xlsx",
                "comparison.xlsx",
                "Sheet1",
                "xl/worksheets/sheet1.xml",
            ),
        )
        self.renderers = dict(
            renderers
            or {
                "statement": write_statement,
                "quotation": write_quotation,
                "comparison": write_comparison,
            }
        )
        self.capacities = {
            "statement": 10,
            "quotation": 11,
            "comparison": 6,
        }

    def generate(
        self,
        source_path: str | Path,
        output_dir: str | Path,
    ) -> GenerationReport:
        """Generate all documents in staging and publish only a valid full set."""

        source = Path(source_path).resolve()
        destination = Path(output_dir).resolve()
        statuses = {spec.document: "not run" for spec in self.templates}

        try:
            self._validate_templates(statuses)
            template_hashes = {
                spec.document: _sha256(spec.template_path)
                for spec in self.templates
            }
            engine = BusinessRuleEngine(self.schema_path, self.business_rules_path)
            invoice = engine.load_invoice(source)
            self._validate_invoice_totals(invoice)
            models = {
                "statement": engine.create_statement(invoice),
                "quotation": engine.create_quotation(invoice),
                "comparison": engine.create_comparison(invoice),
            }
            comparison_markup = Decimal(
                str(models["comparison"].data["document"]["markup"]["rate"])
            )
            self._validate_capacities(
                len(invoice["document"]["items"]),
                statuses,
            )
            totals = self._document_totals(models)

            destination.parent.mkdir(parents=True, exist_ok=True)
            with tempfile.TemporaryDirectory(
                prefix=".rmntc-generation-",
                dir=destination.parent,
            ) as temporary:
                working_root = Path(temporary)
                staged_dir = working_root / "staged"
                staged_dir.mkdir()

                for spec in self.templates:
                    staged_output = staged_dir / spec.output_name
                    try:
                        self.renderers[spec.document](
                            spec.template_path,
                            staged_output,
                            models[spec.document],
                        )
                        statuses[spec.document] = "generated"
                        self._validate_workbook(
                            spec,
                            staged_output,
                            template_hashes[spec.document],
                            models[spec.document],
                        )
                        statuses[spec.document] = "validated"
                    except Exception:
                        statuses[spec.document] = "failed"
                        raise

                self._promote_all(
                    staged_dir,
                    destination,
                    working_root / "backups",
                )

            results = tuple(
                DocumentResult(
                    document=spec.document,
                    template_path=spec.template_path,
                    output_path=destination / spec.output_name,
                    total=totals[spec.document],
                    status="success",
                )
                for spec in self.templates
            )
            return GenerationReport(
                source_path=source,
                company=self.company,
                comparison_markup=comparison_markup,
                documents=results,
            )
        except Exception as error:
            if isinstance(error, GenerationError):
                if not error.statuses:
                    error.statuses.update(statuses)
                raise
            raise GenerationError(str(error), statuses=statuses) from error

    def configuration(self) -> tuple[Decimal, tuple[TemplateSpec, ...]]:
        """Return configured markup and immutable template selections."""

        engine = BusinessRuleEngine(self.schema_path, self.business_rules_path)
        return engine.pricing.comparison_markup_rate, self.templates

    def _validate_templates(self, statuses: dict[str, str]) -> None:
        for spec in self.templates:
            if not spec.template_path.is_file():
                statuses[spec.document] = "failed"
                raise GenerationError(
                    f"Missing {spec.document} template: {spec.template_path}",
                    statuses=statuses,
                )
            if not zipfile.is_zipfile(spec.template_path):
                statuses[spec.document] = "failed"
                raise GenerationError(
                    f"Invalid XLSX template for {spec.document}: "
                    f"{spec.template_path}",
                    statuses=statuses,
                )

    @staticmethod
    def _validate_invoice_totals(invoice: Mapping[str, object]) -> None:
        document = invoice["document"]
        if not isinstance(document, Mapping):
            raise GenerationError("Canonical invoice document must be an object.")
        items = document["items"]
        totals = document["totals"]
        if not isinstance(items, list) or not isinstance(totals, Mapping):
            raise GenerationError("Canonical invoice items or totals are invalid.")
        for field in ("supply_amount", "vat", "total"):
            line_sum = sum(Decimal(str(item[field])) for item in items)
            declared = Decimal(str(totals[field]))
            if line_sum != declared:
                raise GenerationError(
                    f"Canonical invoice {field} total {declared} does not equal "
                    f"the line-item sum {line_sum}."
                )

    def _validate_capacities(
        self,
        item_count: int,
        statuses: dict[str, str],
    ) -> None:
        exceeded = {
            document: capacity
            for document, capacity in self.capacities.items()
            if item_count > capacity
        }
        if not exceeded:
            return
        for document in exceeded:
            statuses[document] = "failed"
        details = ", ".join(
            f"{document} supports {capacity}"
            for document, capacity in exceeded.items()
        )
        raise GenerationError(
            f"Invoice has {item_count} items; template capacity exceeded "
            f"({details}).",
            statuses=statuses,
        )

    @staticmethod
    def _document_totals(models: Mapping[str, ViewModel]) -> dict[str, Decimal]:
        statement = models["statement"].data["document"]["totals"]["total"]
        quotation = models["quotation"].data["document"]["totals"]["total"]
        comparison = models["comparison"].data["document"]["totals"][
            "comparison"
        ]["supply_amount"]
        return {
            "statement": Decimal(str(statement)),
            "quotation": Decimal(str(quotation)),
            "comparison": Decimal(str(comparison)),
        }

    def _validate_workbook(
        self,
        spec: TemplateSpec,
        generated_path: Path,
        source_hash: str,
        view_model: ViewModel,
    ) -> None:
        if not generated_path.is_file():
            raise GenerationError(
                f"{spec.document} renderer did not create {generated_path}."
            )
        if _sha256(spec.template_path) != source_hash:
            raise GenerationError(
                f"{spec.document} source template changed during generation."
            )
        source_parts = _xlsx_parts(spec.template_path)
        generated_parts = _xlsx_parts(generated_path)
        if source_parts.keys() != generated_parts.keys():
            raise GenerationError(
                f"{spec.document} output changed the OOXML package member set."
            )
        for name, payload in source_parts.items():
            if name != spec.sheet_part and generated_parts[name] != payload:
                raise GenerationError(
                    f"{spec.document} output changed preserved OOXML part {name}."
                )

        before_sheet = source_parts[spec.sheet_part]
        after_sheet = generated_parts[spec.sheet_part]
        original_formulas = _formula_map(before_sheet)
        generated_formulas = _formula_map(after_sheet)
        if any(
            generated_formulas.get(cell) != formula
            for cell, formula in original_formulas.items()
            if not (
                spec.document == "statement" and cell in {"D2", "E18"}
            )
        ):
            raise GenerationError(
                f"{spec.document} output changed template formulas."
            )
        if _merge_refs(before_sheet) != _merge_refs(after_sheet):
            raise GenerationError(
                f"{spec.document} output changed merged ranges."
            )
        if _cell_style_map(before_sheet) != _cell_style_map(after_sheet):
            raise GenerationError(
                f"{spec.document} output changed mapped cell styles."
            )
        if _worksheet_structure(before_sheet) != _worksheet_structure(after_sheet):
            raise GenerationError(
                f"{spec.document} output changed dimensions or print/page settings."
            )

        workbook = load_workbook(generated_path, data_only=False, read_only=False)
        try:
            if spec.worksheet_name not in workbook.sheetnames:
                raise GenerationError(
                    f"{spec.document} output is missing worksheet "
                    f"{spec.worksheet_name!r}."
                )
            self._validate_rendered_values(
                spec.document,
                workbook[spec.worksheet_name],
                view_model,
            )
        finally:
            workbook.close()
        if spec.document in {"quotation", "comparison"}:
            calculated = load_workbook(
                generated_path,
                data_only=True,
                read_only=False,
            )
            try:
                sheet = calculated[spec.worksheet_name]
                if spec.document == "quotation":
                    expected = Decimal(
                        str(view_model.data["document"]["totals"]["total"])
                    )
                    actual = Decimal(str(sheet["F29"].value))
                else:
                    expected = Decimal(
                        str(
                            view_model.data["document"]["totals"]["comparison"][
                                "supply_amount"
                            ]
                        )
                    )
                    actual = Decimal(str(sheet["G20"].value))
                if actual != expected:
                    raise GenerationError(
                        f"{spec.document} cached total {actual} does not equal "
                        f"the expected total {expected}."
                    )
            finally:
                calculated.close()

    @staticmethod
    def _validate_rendered_values(document: str, sheet, model: ViewModel) -> None:
        data = model.data["document"]
        items = data["items"]
        if document == "statement":
            rows = (7, 8, 10, 11, 12, 13, 14, 15, 16, 17)
            for row, item in zip(rows, items):
                if (
                    sheet[f"B{row}"].value != item["quantity"]
                    or sheet[f"C{row}"].value != item["description"]
                    or sheet[f"D{row}"].value != item["unit_price"]
                ):
                    raise GenerationError("Statement item values failed validation.")
            expected_total = data["totals"]["total"]
            if sheet["D2"].value != expected_total or sheet["E18"].value != expected_total:
                raise GenerationError("Statement VAT-included total failed validation.")
        elif document == "quotation":
            for row, item in zip(range(16, 27), items):
                if (
                    sheet[f"B{row}"].value != item["description"]
                    or sheet[f"D{row}"].value != item["quantity"]
                    or sheet[f"E{row}"].value != item["unit_price"]
                ):
                    raise GenerationError("Quotation item values failed validation.")
            if (
                sheet["F27"].value != "=SUM(F16:G26)"
                or sheet["F28"].value != data["totals"]["vat"]
                or sheet["F29"].value != "=F27+F28"
            ):
                raise GenerationError("Quotation totals failed validation.")
        else:
            for row, item in zip(range(14, 20), items):
                if (
                    sheet[f"A{row}"].value != item["description"]
                    or sheet[f"C{row}"].value != item["quantity"]
                    or sheet[f"E{row}"].value
                    != item["comparison"]["unit_price"]
                ):
                    raise GenerationError("Comparison values failed validation.")
            expected = sum(
                Decimal(str(item["comparison"]["supply_amount"]))
                for item in items
            )
            declared = Decimal(
                str(data["totals"]["comparison"]["supply_amount"])
            )
            if declared != expected or sheet["G20"].value != "=SUM(G14:G19)":
                raise GenerationError("Comparison totals failed validation.")

    def _promote_all(
        self,
        staged_dir: Path,
        destination: Path,
        backup_dir: Path,
    ) -> None:
        destination.mkdir(parents=True, exist_ok=True)
        backup_dir.mkdir()
        promoted: list[Path] = []
        backups: dict[Path, Path] = {}
        try:
            for spec in self.templates:
                final_path = destination / spec.output_name
                if final_path.exists():
                    backup_path = backup_dir / spec.output_name
                    os.replace(final_path, backup_path)
                    backups[final_path] = backup_path
            for spec in self.templates:
                final_path = destination / spec.output_name
                os.replace(staged_dir / spec.output_name, final_path)
                promoted.append(final_path)
        except Exception as error:
            for path in promoted:
                if path.exists():
                    path.unlink()
            for final_path, backup_path in backups.items():
                if backup_path.exists():
                    os.replace(backup_path, final_path)
            raise GenerationError(
                f"Could not atomically publish generated workbooks: {error}"
            ) from error


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _xlsx_parts(path: Path) -> dict[str, bytes]:
    try:
        with zipfile.ZipFile(path) as archive:
            bad_member = archive.testzip()
            if bad_member:
                raise GenerationError(
                    f"Workbook {path} failed ZIP integrity at {bad_member}."
                )
            return {name: archive.read(name) for name in archive.namelist()}
    except (OSError, zipfile.BadZipFile) as error:
        raise GenerationError(f"Workbook {path} cannot be opened: {error}") from error


def _formula_map(xml: bytes) -> dict[str, tuple[str, tuple[tuple[str, str], ...]]]:
    root = ElementTree.fromstring(xml)
    return {
        cell.attrib["r"]: (
            formula.text or "",
            tuple(sorted(formula.attrib.items())),
        )
        for cell in root.findall(".//x:c", WORKSHEET_NS)
        if (formula := cell.find("x:f", WORKSHEET_NS)) is not None
    }


def _merge_refs(xml: bytes) -> tuple[str, ...]:
    root = ElementTree.fromstring(xml)
    return tuple(
        merge.attrib["ref"]
        for merge in root.findall(".//x:mergeCell", WORKSHEET_NS)
    )


def _cell_style_map(xml: bytes) -> dict[str, str | None]:
    root = ElementTree.fromstring(xml)
    return {
        cell.attrib["r"]: cell.attrib.get("s")
        for cell in root.findall(".//x:c", WORKSHEET_NS)
    }


def _worksheet_structure(xml: bytes) -> dict[bytes, tuple[bytes, ...]]:
    result: dict[bytes, tuple[bytes, ...]] = {}
    for tag in (
        b"cols",
        b"dimension",
        b"pageMargins",
        b"pageSetup",
        b"printOptions",
        b"sheetPr",
        b"sheetViews",
    ):
        pattern = re.compile(
            rb"<" + tag + rb"(?:\s[^>]*)?(?:/>|>.*?</" + tag + rb">)",
            re.DOTALL,
        )
        result[tag] = tuple(pattern.findall(xml))
    result[b"row"] = tuple(re.findall(rb"<row(?:\s[^>]*)?>", xml))
    return result
